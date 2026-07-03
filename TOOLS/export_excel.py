#!/usr/bin/env python3
"""Export the Bhagavad Gita corpus to an Excel workbook (gita.xlsx).

Reads gita.db (built by build_db.py) and writes three sheets:
    Chapters      - one row per chapter
    Sloks         - one row per verse (core fields + image-generation prompt)
    Commentaries  - one row per translation/commentary (slok x author x field)

The Sloks sheet includes a `prompt` column: a complete, self-contained
image-generation JSON for each verse. The constant base of that prompt
(art_style / style_rules / negative_prompt / output_format) is parsed from
.claude/Bhagavad_Gita_Image_Generation_Guidelines.md so that file stays the
single source of truth; per-verse details (chapter, speaker, Sanskrit text,
English meaning, etc.) are merged in from gita.db.

Excel limits a single cell to 32,767 characters; longer commentaries are
truncated with a marker (the full text always remains in gita.db). Run:

    python3 TOOLS/export_excel.py
"""
import copy
import json
import re
import sqlite3
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

CELL_LIMIT = 32_767

GUIDELINES_REL = ".claude/Bhagavad_Gita_Image_Generation_Guidelines.md"

# Order of preference for the plain English meaning fed to the image model.
ENGLISH_PRIORITY = ["prabhu", "purohit", "siva", "gambir", "adi", "raman", "san"]
# Texts that are placeholders, not real translations of the verse.
PLACEHOLDER_MARKERS = (
    "did not comment", "commentary starts from", "no commentary",
)


def clean(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if len(value) > CELL_LIMIT:
            marker = " …[truncated; full text in gita.db]"
            value = value[: CELL_LIMIT - len(marker)] + marker
            clean.truncated += 1
    return value


clean.truncated = 0


def load_base_prompt(md_path):
    """Parse the ```json blocks from the guidelines markdown and merge them
    into a single base-prompt dict (art_style, style_rules, negative_prompt,
    output_format)."""
    if not md_path.exists():
        die(f"guidelines file not found: {md_path}")
    text = md_path.read_text(encoding="utf-8")
    blocks = re.findall(r"```\s*json\s*(.*?)```", text, flags=re.DOTALL)
    if not blocks:
        die(f"no ```json blocks found in {md_path}")
    base = {}
    for block in blocks:
        try:
            base.update(json.loads(block))
        except json.JSONDecodeError as exc:
            die(f"invalid JSON block in {md_path}: {exc}")
    missing = {"art_style", "style_rules", "negative_prompt", "output_format"} - set(base)
    if missing:
        die(f"guidelines missing expected sections: {sorted(missing)}")
    return base


def die(msg):
    print(f"ERROR: {msg}", file=sys.stderr)
    sys.exit(1)


def pick_english_meaning(et_by_author):
    """Choose the best real English translation from {author_code: text}."""
    for code in ENGLISH_PRIORITY:
        txt = et_by_author.get(code)
        if txt and not any(m in txt.lower() for m in PLACEHOLDER_MARKERS):
            return txt.strip()
    # Fall back to any non-placeholder English text available.
    for txt in et_by_author.values():
        if txt and not any(m in txt.lower() for m in PLACEHOLDER_MARKERS):
            return txt.strip()
    return None


def output_format_with_names(base_output_format, stem):
    """Deep-copy the shared output_format and give each image variant its own
    file_name: <stem>_<variant> where variant is square/portrait/landscape."""
    of = copy.deepcopy(base_output_format)
    for img_key, spec in of.get("images", {}).items():
        variant = img_key.split("_")[0]  # portrait_9_16 -> portrait, etc.
        of["images"][img_key] = {"file_name": f"{stem}_{variant}", **spec}
    return of


def build_chapter_prompt(base, chapter):
    """Assemble a chapter cover-image prompt JSON (as a string)."""
    stem = f"chapter_{chapter['chapter_number']}"
    prompt = {
        "task": (
            "You are a master classical Indian painter creating the cover "
            "illustration for this chapter of the Bhagavad Gita. Capture the "
            "chapter's overarching theme in ONE unified scene and express it as "
            "detailed image prompts following output_format. Strictly obey "
            "art_style, style_rules and negative_prompt so every illustration "
            "looks like it belongs in the same illuminated ancient Indian "
            "manuscript."
        ),
        "chapter": {
            "file_name": stem,
            "chapter": chapter["chapter_number"],
            "name": chapter["name"],
            "title": chapter["translation"],
            "transliteration": chapter["transliteration"],
            "meaning": chapter["meaning_en"],
            "summary": chapter["summary_en"],
            "verses_count": chapter["verses_count"],
        },
        "art_style": base["art_style"],
        "style_rules": base["style_rules"],
        "negative_prompt": base["negative_prompt"],
        "output_format": output_format_with_names(base["output_format"], stem),
    }
    return json.dumps(prompt, ensure_ascii=False, indent=2)


def build_prompt(base, slok, chapter, english_meaning):
    """Assemble the full per-verse image-generation prompt JSON (as a string)."""
    stem = f"chapter_{slok['chapter_number']}_slok_{slok['verse']}"
    prompt = {
        "task": (
            "You are a master classical Indian painter creating an illustration "
            "for this Bhagavad Gita shloka. First understand the verse's meaning, "
            "then design ONE unified scene concept and express it as detailed image "
            "prompts following output_format. Strictly obey art_style, style_rules "
            "and negative_prompt so every illustration looks like it belongs in the "
            "same illuminated ancient Indian manuscript."
        ),
        "shloka": {
            "file_name": stem,
            "verse_id": slok["id"],
            "chapter": slok["chapter_number"],
            "chapter_title": chapter["translation"],
            "chapter_meaning": chapter["meaning_en"],
            "verse": slok["verse"],
            "speaker": slok["speaker"],
            "sanskrit": slok["slok_text"],
            "transliteration": slok["transliteration"],
            "english_meaning": english_meaning,
        },
        "art_style": base["art_style"],
        "style_rules": base["style_rules"],
        "negative_prompt": base["negative_prompt"],
        "output_format": output_format_with_names(base["output_format"], stem),
    }
    return json.dumps(prompt, ensure_ascii=False, indent=2)


def write_sheet(ws, headers, rows, widths=None):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append([clean(v) for v in row])
    for i, _ in enumerate(headers, start=1):
        w = (widths or {}).get(i, 22)
        ws.column_dimensions[get_column_letter(i)].width = w


def main():
    root = Path(__file__).resolve().parent.parent
    db_path = root / "gita.db"
    out_path = root / "gita.xlsx"
    if not db_path.exists():
        die("gita.db not found - run TOOLS/build_db.py first")

    base = load_base_prompt(root / GUIDELINES_REL)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    wb = Workbook()

    # Chapters ---------------------------------------------------------------
    ws = wb.active
    ws.title = "Chapters"
    chapter_rows = conn.execute(
        "SELECT chapter_number, name, translation, transliteration, "
        "meaning_en, meaning_hi, summary_en, summary_hi, verses_count "
        "FROM chapters ORDER BY chapter_number"
    ).fetchall()
    chapters = {r["chapter_number"]: r for r in chapter_rows}
    write_sheet(ws, [
        "chapter", "name", "translation", "transliteration",
        "meaning_en", "meaning_hi", "summary_en", "summary_hi", "verses_count",
        "prompt",
    ], [tuple(r) + (build_chapter_prompt(base, r),) for r in chapter_rows],
        widths={10: 80})
    n_ch = len(chapter_rows)

    # English meaning per slok ----------------------------------------------
    et_map = {}  # slok_id -> {author_code: text}
    for r in conn.execute(
        "SELECT slok_id, author_code, text FROM translations WHERE field_type='et'"
    ):
        et_map.setdefault(r["slok_id"], {})[r["author_code"]] = r["text"]

    # Sloks (with prompt column) --------------------------------------------
    ws = wb.create_sheet("Sloks")
    slok_rows = conn.execute(
        "SELECT s.id, s.chapter_number, s.verse, sp.name AS speaker, "
        "s.slok_text, s.transliteration, s.image_url "
        "FROM sloks s LEFT JOIN speakers sp ON sp.speaker_id = s.speaker_id "
        "ORDER BY s.chapter_number, s.verse"
    ).fetchall()

    out_rows = []
    no_meaning = 0
    for s in slok_rows:
        meaning = pick_english_meaning(et_map.get(s["id"], {}))
        if meaning is None:
            no_meaning += 1
        prompt = build_prompt(base, s, chapters[s["chapter_number"]], meaning)
        out_rows.append((
            s["id"], s["chapter_number"], s["verse"], s["speaker"],
            s["slok_text"], s["transliteration"], s["image_url"], prompt,
        ))
    write_sheet(ws, [
        "id", "chapter", "verse", "speaker", "slok", "transliteration",
        "image_url", "prompt",
    ], out_rows, widths={8: 80})
    n_sl = len(out_rows)

    # Commentaries / translations -------------------------------------------
    ws = wb.create_sheet("Commentaries")
    comm_rows = conn.execute(
        "SELECT t.slok_id, s.chapter_number, s.verse, t.author_code, a.name, "
        "t.field_type, t.ordinal, t.text "
        "FROM translations t "
        "JOIN sloks s ON s.id = t.slok_id "
        "JOIN authors a ON a.author_code = t.author_code "
        "ORDER BY s.chapter_number, s.verse, t.ordinal"
    ).fetchall()
    write_sheet(ws, [
        "slok_id", "chapter", "verse", "author_code", "author_name",
        "field_type", "ordinal", "text",
    ], [tuple(r) for r in comm_rows])
    n_tr = len(comm_rows)

    conn.close()
    wb.save(out_path)

    size = out_path.stat().st_size
    print(f"Wrote {out_path.name} ({size/1024/1024:.1f} MB)")
    print(f"  Chapters     : {n_ch} rows")
    print(f"  Sloks        : {n_sl} rows (with prompt column)")
    print(f"  Commentaries : {n_tr} rows")
    if no_meaning:
        print(f"  note: {no_meaning} slok(s) had no usable English meaning "
              f"(english_meaning set to null in prompt)")
    if clean.truncated:
        print(f"  note: {clean.truncated} cell(s) exceeded Excel's 32,767-char limit "
              f"and were truncated (full text remains in gita.db)")


if __name__ == "__main__":
    main()
